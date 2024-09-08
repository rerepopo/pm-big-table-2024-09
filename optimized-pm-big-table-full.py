from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class PMBigTable:
    def __init__(self, working_days_per_week=5, project_start_date=datetime(2024, 1, 1), duration_unit='day', display_unit='day', holidays=None):
        self.working_days_per_week = working_days_per_week
        self.project_start_date = project_start_date
        self.duration_unit = duration_unit
        self.display_unit = display_unit
        self.holidays = holidays or []
        self.wbs_data = []
        self.task_schedule = {}
        self.owner_schedules = {}
        self.critical_path = []

    def is_working_day(self, date):
        return (date.weekday() < self.working_days_per_week and date.date() not in self.holidays)

    def next_working_day(self, date):
        next_day = date + timedelta(days=1)
        while not self.is_working_day(next_day):
            next_day += timedelta(days=1)
        return next_day

    def load_wbs(self, wbs_data):
        self.wbs_data = sorted(wbs_data, key=lambda x: int(x['WBS']))
        self._calculate_schedule()
        self._calculate_critical_path()

    def _calculate_schedule(self):
        # Calculate schedule and allocate tasks
        for task in self.wbs_data:
            predecessors = task['Predecessors'].split(',') if task['Predecessors'] else []
            pred_end_dates = [self.task_schedule[pred.strip()]['end'] for pred in predecessors if pred.strip()]
            start_date = max(pred_end_dates) if pred_end_dates else self.project_start_date
            start_date = self._find_next_available_date(start_date, task['Owner'])
            
            duration = self._convert_duration(task['Duration'])
            owner = task['Owner']
            
            if owner not in self.owner_schedules:
                self.owner_schedules[owner] = {}
            
            end_date = self._allocate_task(task['Task'], start_date, duration, owner)
            
            self.task_schedule[task['Task']] = {
                'start': start_date,
                'end': end_date,
                'owner': owner,
                'duration': duration
            }

    def _find_next_available_date(self, start_date, owner):
        current_date = start_date
        while True:
            if self.is_working_day(current_date):
                if owner not in self.owner_schedules:
                    self.owner_schedules[owner] = {}
                if current_date not in self.owner_schedules[owner] or self.owner_schedules[owner][current_date] < 8:
                    return current_date
            current_date = self.next_working_day(current_date)

    def _allocate_task(self, task_name, start_date, duration, owner):
        remaining_hours = duration * 8
        current_date = start_date
        while remaining_hours > 0:
            if self.is_working_day(current_date):
                if current_date not in self.owner_schedules[owner]:
                    self.owner_schedules[owner][current_date] = 0
                available_hours = min(8 - self.owner_schedules[owner][current_date], remaining_hours)
                self.owner_schedules[owner][current_date] += available_hours
                remaining_hours -= available_hours
            if remaining_hours > 0:
                current_date = self.next_working_day(current_date)
        return current_date

    def _convert_duration(self, duration):
        return duration / 8 if self.duration_unit == 'hour' else duration

    def _calculate_critical_path(self):
        # Calculate critical path based on resource constraints
        sorted_tasks = sorted(self.wbs_data, key=lambda x: self.task_schedule[x['Task']]['start'])
        current_time = self.project_start_date
        self.critical_path = []
        
        for task in sorted_tasks:
            task_name = task['Task']
            task_info = self.task_schedule[task_name]
            self.critical_path.append(task_name)
            task_info['ES'] = current_time
            task_info['EF'] = task_info['end']
            task_info['LS'] = task_info['ES']
            task_info['LF'] = task_info['EF']
            task_info['Float'] = 0
            current_time = task_info['EF']

    def print_schedule(self):
        print("\nProject Schedule:")
        for task in self.wbs_data:
            task_name = task['Task']
            schedule = self.task_schedule[task_name]
            print(f"{task_name}: Start - {schedule['start'].date()}, End - {schedule['end'].date()}, Duration - {schedule['duration']} days, Owner - {schedule['owner']}")

    def print_gantt_chart(self):
        print("\nGantt Chart:")
        timeline = self._generate_timeline()
        print(timeline)
        for task in self.wbs_data:
            task_name = task['Task']
            start = self.task_schedule[task_name]['start']
            end = self.task_schedule[task_name]['end']
            chart_line = f"{task_name:<10}"
            for date in self._date_range(self.project_start_date, max(task['end'] for task in self.task_schedule.values())):
                if start <= date <= end:
                    chart_line += "=" if self.is_working_day(date) and self.owner_schedules[task['Owner']].get(date, 0) > 0 else "-"
                else:
                    chart_line += " "
            chart_line += f" {task['Owner']}"
            print(chart_line)

    def _generate_timeline(self):
        timeline = "          "
        last_end_date = max(task['end'] for task in self.task_schedule.values())
        for date in self._date_range(self.project_start_date, last_end_date):
            timeline += str(date.day)[0] if self.display_unit == 'day' else (str(date.isocalendar()[1] % 10) if date.weekday() == 0 else " ")
        return timeline

    def _date_range(self, start_date, end_date):
        for n in range(int((end_date - start_date).days) + 1):
            yield start_date + timedelta(n)

    def print_critical_path(self):
        print("\nCritical Path (Resource Constrained):")
        total_duration = sum(self.task_schedule[task]['duration'] for task in self.critical_path)
        for task in self.critical_path:
            task_info = self.task_schedule[task]
            print(f"  {task}: {task_info['duration']} days - Start: {task_info['ES'].date()}, End: {task_info['EF'].date()}, Owner: {task_info['owner']}")
        print(f"Total duration of critical path: {total_duration} days")

        print("\nAll Tasks Schedule:")
        for task_name, task_info in self.task_schedule.items():
            print(f"  {task_name}: ES={task_info['ES'].date()}, EF={task_info['EF'].date()}, "
                  f"LS={task_info['LS'].date()}, LF={task_info['LF'].date()}, "
                  f"Float={task_info['Float']}, Duration={task_info['duration']}")

    def print_detailed_schedule(self):
        print("\nDetailed Schedule:")
        for task in self.wbs_data:
            task_name = task['Task']
            info = self.task_schedule[task_name]
            print(f"{task_name}:")
            print(f"  Start: {info['start'].date()}")
            print(f"  End: {info['end'].date()}")
            print(f"  Duration: {info['duration']} days")
            print(f"  Owner: {info['owner']}")
            print("  Daily work hours:")
            for date, hours in self.owner_schedules[info['owner']].items():
                if info['start'] <= date <= info['end']:
                    print(f"    {date.date()}: {hours} hours")
            print()

    def generate_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Schedule"
        
        # Add headers
        headers = ['WBS ID', 'Task Name', 'Owner', 'Planned Start', 'Planned End', 'Planned Hours', 'Actual Hours', 'Progress', 'Earned Value (EV)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add task data
        for row, task in enumerate(self.wbs_data, start=2):
            ws.cell(row=row, column=1, value=task['WBS'])
            ws.cell(row=row, column=2, value=task['Task'])
            ws.cell(row=row, column=3, value=task['Owner'])
            ws.cell(row=row, column=4, value=self.task_schedule[task['Task']]['start'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=self.task_schedule[task['Task']]['end'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=task['Duration'] * 8)
            ws.cell(row=row, column=8, value=0).number_format = '0.00%'
            ws.cell(row=row, column=9, value=f"=F{row}*H{row}")
        
        # Add totals
        total_row = len(self.wbs_data) + 2
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")
        ws.cell(row=total_row, column=8, value=f"=I{total_row}/F{total_row}").number_format = '0.00%'
        ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{total_row-1})")
        
        # Add EVM rows
        evm_start_row = total_row + 2
        evm_rows = ['Planned Value (PV)', 'Cumulative PV', 'Actual Cost (AC)', 'Cumulative AC', 'Cumulative EV (Manual)']
        for i, row_title in enumerate(evm_rows):
            ws.cell(row=evm_start_row + i, column=1, value=row_title)
        
        # Add daily data
        project_end = max(task['end'] for task in self.task_schedule.values())
        date_range = [d for d in self._date_range(self.project_start_date, project_end) if self.is_working_day(d)]
        
        for col, date in enumerate(date_range, start=10):
            col_letter = get_column_letter(col)
            ws.cell(row=1, column=col, value=date.strftime('%m/%d'))
            ws.column_dimensions[col_letter].width = 5.2
            
            # Add week number and alternate shading
            week_number = date.isocalendar()[1]
            fill_color = "F0F0F0" if week_number % 2 == 1 else "E0E0E0"
            for row in range(1, total_row):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row=total_row, column=col, value=f"W{week_number:02d}")
            
            # Add task hours
            for row, task in enumerate(self.wbs_data, start=2):
                if self.task_schedule[task['Task']]['start'] <= date <= self.task_schedule[task['Task']]['end']:
                    cell = ws.cell(row=row, column=col)
                    cell.value = "8\n0"
                    cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add EVM formulas
            pv_formula = f'=SUM(IF({col_letter}2:{col_letter}{total_row-1}="",0,IF(ISERROR(FIND(CHAR(10),{col_letter}2:{col_letter}{total_row-1})),{col_letter}2:{col_letter}{total_row-1},VALUE(LEFT({col_letter}2:{col_letter}{total_row-1},FIND(CHAR(10),{col_letter}2:{col_letter}{total_row-1}&CHAR(10))-1)))))'
            ac_formula = f'=SUM(IF({col_letter}2:{col_letter}{total_row-1}="",0,IF(ISERROR(FIND(CHAR(10),{col_letter}2:{col_letter}{total_row-1})),0,VALUE(MID({col_letter}2:{col_letter}{total_row-1},FIND(CHAR(10),{col_letter}2:{col_letter}{total_row-1})+1,LEN({col_letter}2:{col_letter}{total_row-1}))))))'
            
            ws.cell(row=evm_start_row, column=col, value=pv_formula)
            ws.cell(row=evm_start_row + 1, column=col, value=f"=SUM({get_column_letter(10)}{evm_start_row}:{col_letter}{evm_start_row})")
            ws.cell(row=evm_start_row + 2, column=col, value=ac_formula)
            ws.cell(row=evm_start_row + 3, column=col, value=f"=SUM({get_column_letter(10)}{evm_start_row+2}:{col_letter}{evm_start_row+2})")
        
        # Add actual hours formula
        last_date_column = get_column_letter(9 + len(date_range))
        for row in range(2, total_row):
            actual_hours_formula = f'=SUM(IF(J{row}:{last_date_column}{row}="", 0, IF(ISERROR(FIND(CHAR(10), J{row}:{last_date_column}{row})), 0, VALUE(MID(J{row}:{last_date_column}{row}, FIND(CHAR(10), J{row}:{last_date_column}{row}) + 1, LEN(J{row}:{last_date_column}{row}) - FIND(CHAR(10), J{row}:{last_date_column}{row}))))))'
            ws.cell(row=row, column=7, value=actual_hours_formula)

        # Add warning message
        warning_row = evm_start_row + 6
        warning_message = "IMPORTANT: After opening the Excel file for the first time, please use the 'Find and Replace' function to remove all '@' symbols from the document."
        ws.cell(row=warning_row, column=1, value=warning_message)
        ws.merge_cells(start_row=warning_row, start_column=1, end_row=warning_row, end_column=9)
        ws.cell(row=warning_row, column=1).font = Font(bold=True, color="FF0000")
        
        wb.save(filename)
        print(f"Excel file '{filename}' has been generated.")

# Example usage
if __name__ == "__main__":
    holidays = [datetime(2024, 2, 8).date(),datetime(2024, 2, 9).date(),datetime(2024, 2, 12).date(),datetime(2024, 2, 13).date(),datetime(2024, 2, 14).date()]  # Example holidays
    pm_tool = PMBigTable(working_days_per_week=5, project_start_date=datetime(2024, 1, 2), duration_unit='day', display_unit='day', holidays=holidays)
    wbs_data = [
        {'WBS': '1', 'Task': 'Task A', 'Duration': 2, 'Predecessors': '', 'Owner': 'Alex'},
        {'WBS': '2', 'Task': 'Task B', 'Duration': 3, 'Predecessors': '', 'Owner': 'Bob'},
        {'WBS': '3', 'Task': 'Task C', 'Duration': 5, 'Predecessors': '', 'Owner': 'Alex'},
        {'WBS': '4', 'Task': 'Task D', 'Duration': 2, 'Predecessors': '', 'Owner': 'David'},
        {'WBS': '5', 'Task': 'Task E', 'Duration': 3, 'Predecessors': 'Task D', 'Owner': 'Alex'},
        {'WBS': '6', 'Task': 'Task F', 'Duration': 4, 'Predecessors': 'Task D', 'Owner': 'Alex'},
        {'WBS': '7', 'Task': 'Task G', 'Duration': 1, 'Predecessors': 'Task B, Task F', 'Owner': 'Alex'},
        {'WBS': '8', 'Task': 'Task H', 'Duration': 20, 'Predecessors': 'Task G, Task E', 'Owner': 'Alex'},
        {'WBS': '9', 'Task': 'Task M', 'Duration': 6, 'Predecessors': 'Task H', 'Owner': 'Alex'},
    ]
    pm_tool.load_wbs(wbs_data)
    pm_tool.print_schedule()
    pm_tool.print_gantt_chart()
    pm_tool.print_critical_path()
    pm_tool.print_detailed_schedule()
    pm_tool.generate_excel("project_schedule.xlsx")
